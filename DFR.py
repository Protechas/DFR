import sys
import os
import csv
import time
import datetime
import threading
import openpyxl
from openpyxl.styles.alignment import Alignment
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QLabel, QTextEdit, QFileDialog, QProgressBar
from PyQt5.QtWidgets import QVBoxLayout, QHBoxLayout, QSizePolicy, QSpacerItem
from PyQt5.QtGui import QColor, QPalette, QFont, QTextOption
from PyQt5.QtCore import Qt, QObject, pyqtSignal

class CSVEditor(QWidget):
    progress_update = pyqtSignal(int, int)

    def __init__(self):
        super().__init__()
        self.current_theme = "dark"  # Initial theme
        self.light_style = '''
            QWidget {
                background-color: #eee;
                color: #222;
            }
 
            QLabel {
                color: #000000;  /* Dark blue text */
            }
 
            QPushButton {
                background-color: #0066cc;  /* Dark blue background */
                color: #eee;  /* Light text */
                border: 2px solid #0066cc;  /* Dark blue border */
                border-radius: 10px;  /* Border radius for a "pop" effect */
                padding: 10px;  /* Increased padding for a "pop" effect */
                margin: 5px;
            }
 
            QPushButton:hover {
                background-color: #004080;  /* Darker blue on hover */
                border: 2px solid #004080;  /* Darker blue border on hover */
            }
        '''
 
        self.dark_style = '''
            QWidget {
                background-color: #222;
                color: #eee;
            }
 
            QLabel {
                color: #008080;  /* Light blue text */
            }
 
            QPushButton {
                background-color: #66ccff;  /* Light blue background */
                color: #222;  /* Dark text */
                border: 2px solid #66ccff;  /* Light blue border */
                border-radius: 10px;  /* Border radius for a "pop" effect */
                padding: 10px;  /* Increased padding for a "pop" effect */
                margin: 5px;
            }
 
            QPushButton:hover {
                background-color: #3385ff;  /* Lighter blue on hover */
                border: 2px solid #3385ff;  /* Lighter blue border on hover */
            }
        '''

        self.initUI()

    def initUI(self):
        self.setWindowTitle("DFR")
        self.center()

        layout = QVBoxLayout()

        # Title Label
        self.title_label = QLabel("Attach the CSV File needing Editing")
        self.title_label.setFont(QFont("Arial", 14))
        layout.addWidget(self.title_label)

        # File Selection Section
        file_layout = QHBoxLayout()
        self.select_button = QPushButton("Select File")
        self.select_button.setStyleSheet("background-color: #008080; color: white;")
        self.select_button.clicked.connect(self.select_file)
        file_layout.addWidget(self.select_button)
        self.file_display_text = QTextEdit()
        self.file_display_text.setFixedHeight(30)
        self.file_display_text.setReadOnly(True)
        file_layout.addWidget(self.file_display_text)
        layout.addLayout(file_layout)

        # Template Selection Section
        template_layout = QHBoxLayout()
        self.select_template_button = QPushButton("Select Template for Header")
        self.select_template_button.setStyleSheet("background-color: #008080; color: white;")
        self.select_template_button.clicked.connect(self.select_template)
        template_layout.addWidget(self.select_template_button)
        self.template_display_text = QTextEdit()
        self.template_display_text.setFixedHeight(30)
        self.template_display_text.setReadOnly(True)
        template_layout.addWidget(self.template_display_text)
        layout.addLayout(template_layout)

        # Edit Button
        self.edit_button = QPushButton("Edit File")
        self.edit_button.setStyleSheet("background-color: #008080; color: white;")
        self.edit_button.clicked.connect(self.edit_file_wrapper)
        layout.addWidget(self.edit_button)

        # Progress Bar
        self.progress_bar = QProgressBar()
        layout.addWidget(self.progress_bar)

        # Toggle Theme Button
        self.toggle_theme_button = QPushButton("Dark/Light")
        self.toggle_theme_button.clicked.connect(self.toggle_theme)
        self.toggle_theme_button.setStyleSheet("background-color: #008080; color: white;")
        layout.addWidget(self.toggle_theme_button)
        self.setStyleSheet(self.dark_style)
        self.setLayout(layout)

        # Connect progress signal
        self.progress_update.connect(self.update_progress)

    def center(self):
        # Function to center the application window on the screen
        frame_geometry = self.frameGeometry()
        screen = QApplication.desktop().screenNumber(QApplication.desktop().cursor().pos())
        center_point = QApplication.desktop().screenGeometry(screen).center()
        frame_geometry.moveCenter(center_point)
        self.move(frame_geometry.topLeft())

    def set_button_style(self, button):
        button.setStyleSheet(
            "QPushButton { background-color: %s; color: %s; border: none; padding: 10px; border-radius: 5px; }"
            "QPushButton:hover { background-color: %s; color: %s; }"
            % (
                self.color_schemes[self.current_theme]["button_bg"],
                self.color_schemes[self.current_theme]["button_fg"],
                self.color_schemes[self.current_theme]["button_hover_bg"],
                self.color_schemes[self.current_theme]["button_hover_fg"],
            )
        )

    def toggle_theme(self):
        # Function to toggle between dark and light themes
        if self.current_theme == "dark":
            self.setStyleSheet(self.light_style)
            self.current_theme = "light"
        else:
            self.setStyleSheet(self.dark_style)
            self.current_theme = "dark"

    def apply_theme(self):
        theme = self.color_schemes[self.current_theme]
        palette = QPalette()
        palette.setColor(QPalette.Window, QColor(theme["bg"]))
        palette.setColor(QPalette.WindowText, QColor(theme["fg"]))
        self.setPalette(palette)
        for button in [self.select_button, self.select_template_button, self.edit_button, self.toggle_theme_button]:
            self.set_button_style(button)

    def edit_file_thread(self, file_path, template_path=None):
        thread = threading.Thread(target=self.edit_file, args=(file_path, template_path))
        thread.start()

    def select_file(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        file_path, _ = QFileDialog.getOpenFileName(self, "Select File", "", "CSV Files (*.csv);;Excel Files (*.xlsx)", options=options)
        if file_path:
            self.file_display_text.setPlainText(file_path)

    def select_template(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        template_path, _ = QFileDialog.getOpenFileName(self, "Select Template for Header", "", "CSV Files (*.csv);;Excel Files (*.xlsx)", options=options)
        if template_path:
            self.template_display_text.setPlainText(template_path)

    def edit_file_wrapper(self):
        file_path = self.file_display_text.toPlainText()
        template_path = self.template_display_text.toPlainText()
        if file_path:
            self.edit_file_thread(file_path, template_path)

        # Delete the "DFR Month Editing in Progress.xlsx" file if it exists
        progress_file_path = os.path.join(os.path.dirname(file_path), "DFR Month Editing in Progress.xlsx")
        if os.path.exists(progress_file_path):
            os.remove(progress_file_path)

    def update_progress(self, value, max_value):
        progress = int((value / max_value) * 100)
        self.progress_bar.setValue(progress)

    def center_worksheet_cells(self, worksheet):
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

    def hide_columns(self, worksheet, columns_to_hide):
        for col_index in columns_to_hide:
            col_letter = openpyxl.utils.get_column_letter(col_index)
            worksheet.column_dimensions[col_letter].hidden = True

    def edit_file(self, file_path, template_path=None):
        try:
            self.progress_update.emit(1, 5)
            if file_path.endswith('.csv'):
                # Process CSV file
                with open(file_path, 'r', newline='', encoding='utf-8-sig') as csvfile:
                    reader = csv.reader(csvfile)
                    data = list(reader)
            elif file_path.endswith('.xlsx'):
                # Process XLSX file
                workbook = openpyxl.load_workbook(file_path)
                worksheet = workbook.active
                data = []
                for row in worksheet.iter_rows(values_only=True):
                    data.append(list(row))
            else:
                raise ValueError("Unsupported file format")

            for row in data:
                for i in range(len(row)):
                    current_input = row[i]
                    if current_input and current_input in make_mapping:
                        row[i] = make_mapping[current_input]

            for row in data:
                if len(row) > 6:
                    del row[6]

            excel_file_path = os.path.join(os.path.dirname(file_path), "DFR Month Editing in Progress.xlsx")
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            for row_data in data:
                worksheet.append(row_data)
            worksheet.freeze_panes = 'A2'
            self.progress_update.emit(2, 5)
            for row_index in range(worksheet.max_row, 1, -1):
                if worksheet.cell(row=row_index, column=16).value == '':
                    worksheet.delete_rows(row_index)

            edited_ws = workbook.active  # Define edited_ws as the active worksheet
            self.center_worksheet_cells(edited_ws)  # Adjusted to use self.method_name
            self.hide_columns(edited_ws, [1, 2, 3, 4, 5, 10, 11, 12, 13, 14, 21, 30, 31, 32, 33, 34, 35, 36, 37, 38])

            filtered_sheet = workbook.create_sheet(title='False Negatives')
            max_columns = max(len(row) for row in data)
            for row_data in data:
                row_data.extend([''] * (max_columns - len(row_data)))
                if not row_data or row_data[15] == '':
                    filtered_sheet.append(row_data)

            workbook.save(excel_file_path)
            self.progress_update.emit(3, 5)

            if template_path:
                template_wb = openpyxl.load_workbook(template_path)
                template_ws = template_wb.active
                header_row = [cell.value if isinstance(cell, openpyxl.cell.cell.Cell) else cell for cell in template_ws[1]]
                edited_wb = openpyxl.load_workbook(excel_file_path)
                edited_ws = edited_wb.active
                for col_num, value in enumerate(header_row, 1):
                    cell = edited_ws.cell(row=1, column=col_num, value=value)
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='medium'))

                edited_ws.auto_filter.ref = edited_ws.dimensions
                self.sort_sheet(edited_ws)
                self.center_worksheet_cells(edited_ws)  # Corrected function call
                
                # Apply template cell colors to the header only
                for col_index in range(1, template_ws.max_column + 1):
                    template_cell = template_ws.cell(row=1, column=col_index)
                    edited_cell = edited_ws.cell(row=1, column=col_index)
                    if template_cell.fill.start_color.rgb != "00000000":
                        if isinstance(template_cell.fill.start_color.rgb, int):
                            start_color = openpyxl.styles.colors.Color(rgb='%06x' % template_cell.fill.start_color.rgb)
                        else:
                            start_color = template_cell.fill.start_color
                        fill = openpyxl.styles.PatternFill(start_color=start_color, end_color=start_color, fill_type="solid")
                        edited_cell.fill = fill

                current_month = datetime.datetime.now().strftime("%B")
                new_file_name = os.path.join(os.path.dirname(file_path), f"DFR {current_month}.xlsx")
                edited_wb.save(new_file_name)
                print(f"Saved edited file as {new_file_name}")

            self.progress_update.emit(4, 5)
            os.remove(file_path)
            os.remove(excel_file_path)
            print("Deleted original file and edited Excel files.")
        except Exception as e:
            print("An error occurred during editing:")
            print(e)
        finally:
            self.progress_update.emit(5, 5)

    def sort_sheet(self, worksheet):
        def custom_sort(row):
            sort_columns = [8, 6, 15, 16, 20, 21]
            sort_values = [(row[col - 1] if row[col - 1] is not None else '') for col in sort_columns]
            return sort_values
        
        rows_to_sort = list(worksheet.iter_rows(min_row=2, values_only=True))
        new_rows = []
        prev_col_f_value = None
        for row_data in sorted(rows_to_sort, key=custom_sort):
            col_f_value = row_data[5]
            if prev_col_f_value is not None and col_f_value != prev_col_f_value:
                new_separator_row = [''] * len(row_data)
                new_separator_row[5] = 'a'
                new_separator_row[46] = None
                new_rows.append(new_separator_row)
                for col_index in range(1, len(row_data) + 1):
                    if col_index != 47:
                        worksheet.cell(row=len(new_rows) + 1, column=col_index).fill = openpyxl.styles.PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            new_rows.append(row_data)
            prev_col_f_value = col_f_value

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.value = None
        for row_index, row_data in enumerate(new_rows, start=2):
            for col_index, cell_value in enumerate(row_data, start=1):
                cell = worksheet.cell(row=row_index, column=col_index, value=cell_value)
        for col_index in range(39, 48):
            col_letter = openpyxl.utils.get_column_letter(col_index)
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_index}"]
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick'))
                if row_index == worksheet.max_row:
                    cell.border += openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thick'))
            for row_index in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_index}"]
                cell.border += openpyxl.styles.Border(right=openpyxl.styles.Side(style='thick'))
        for col_index in range(40, 46):
            col_letter = openpyxl.utils.get_column_letter(col_index)
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=col_index, max_col=col_index):
                for cell in row:
                    cell.number_format = '0%'

make_mapping = {
    "AMC": "AMC",
    "ACUR": "Acura",
    "AUDI": "Audi",
    "BUIC": "Buick",
    "CADI": "Cadillac",
    "CHEV": "Chevrolet",
    "CHRY": "Chrysler",
    "DODG": "Dodge",
    "FORD": "Ford",
    "GMC": "GMC",
    "HOND": "Honda",
    "HYUN": "Hyundai",
    "INFI": "Infiniti",
    "JAG": "Jaguar",
    "JEEP": "Jeep",
    "KIA" : "Kia",
    "LEXU": "Lexus",
    "LINC": "Lincoln",
    "LR": "Land Rover",
    "MAZD": "Mazda",
    "MERC": "Mercury",
    "MITS": "Mitsubishi",
    "NISS": "Nissan",
    "OLDS": "Oldsmobile",
    "PLYM": "Plymouth",
    "PONT": "Pontiac",
    "RAM": "Ram",
    "SATN": "Saturn",
    "SUBA": "Subaru",
    "TOYO": "Toyota",
    "VW": "Volkswagen"
}

if __name__ == "__main__":
    app = QApplication(sys.argv)
    editor = CSVEditor()
    editor.show()
    sys.exit(app.exec_())
